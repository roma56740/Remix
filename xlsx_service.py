from __future__ import annotations

import io
import re
import sqlite3
import uuid
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from fastapi import UploadFile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import IMPORT_DIR
from database import _connect
from platform_service import _audit
MAX_IMPORT_BYTES = 15 * 1024 * 1024

HEADERS = (
    "user_email",
    "release_id",
    "track_id",
    "date",
    "listens",
    "amount_rub",
    "operation_title",
)

HEADER_ALIASES = {
    "email": "user_email",
    "user": "user_email",
    "пользователь": "user_email",
    "email пользователя": "user_email",
    "релиз": "release_id",
    "id релиза": "release_id",
    "трек": "track_id",
    "id трека": "track_id",
    "дата": "date",
    "прослушивания": "listens",
    "сумма": "amount_rub",
    "сумма руб": "amount_rub",
    "операция": "operation_title",
    "описание": "operation_title",
}


class XLSXImportError(ValueError):
    pass


def ensure_import_directory() -> None:
    IMPORT_DIR.mkdir(parents=True, exist_ok=True)


def _style_workbook(workbook: Workbook) -> None:
    accent = "E37B47"
    dark = "171717"
    white = "FFFFFF"
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True, color=white)
            cell.fill = PatternFill("solid", fgColor=dark)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[1].height = 25
        for column in range(1, sheet.max_column + 1):
            values = [str(sheet.cell(row=row, column=column).value or "") for row in range(1, min(sheet.max_row, 100) + 1)]
            width = min(36, max(12, max((len(value) for value in values), default=10) + 2))
            sheet.column_dimensions[get_column_letter(column)].width = width
        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.sheet_view.showGridLines = False
        sheet.sheet_properties.tabColor = accent


def build_import_template() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Импорт"
    sheet.append(list(HEADERS))
    sheet.append([
        "artist@example.com",
        15,
        42,
        date.today().isoformat(),
        1200,
        350,
        "Начисление за квартал",
    ])
    sheet.append([
        "artist@example.com",
        15,
        "",
        date.today().isoformat(),
        5000,
        "",
        "Прослушивания релиза",
    ])
    info = workbook.create_sheet("Инструкция")
    info.append(["Поле", "Описание"])
    rows = [
        ("user_email", "Email зарегистрированного пользователя. Обязательное поле."),
        ("release_id", "ID релиза. Нужен для импорта прослушиваний релиза."),
        ("track_id", "ID трека. Если указан, прослушивания попадут и в статистику трека."),
        ("date", "Дата статистики в формате ГГГГ-ММ-ДД. Пустое значение означает сегодня."),
        ("listens", "Количество новых прослушиваний за указанную дату. Целое число от 0."),
        ("amount_rub", "Начисление или списание в рублях. Для списания используйте отрицательное число."),
        ("operation_title", "Понятное описание операции, которое увидит пользователь."),
    ]
    for row in rows:
        info.append(row)
    info.append(["Важно", "Строки с ошибками пропускаются. Корректные строки импортируются и сохраняются в истории."])
    _style_workbook(workbook)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_quarter_report() -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Сводка"
    users_sheet = workbook.create_sheet("Пользователи")
    releases_sheet = workbook.create_sheet("Релизы")
    tracks_sheet = workbook.create_sheet("Треки")

    today = date.today()
    quarter = (today.month - 1) // 3 + 1
    start_month = (quarter - 1) * 3 + 1
    start = date(today.year, start_month, 1)
    if start_month + 3 > 12:
        end = date(today.year + 1, 1, 1)
    else:
        end = date(today.year, start_month + 3, 1)

    with _connect() as connection:
        users = connection.execute(
            """
            SELECT u.id, u.name, u.email, u.username, u.balance, u.pending_balance,
                   COUNT(DISTINCT r.id) AS releases,
                   COALESCE((SELECT SUM(l.count) FROM listens l WHERE l.user_id = u.id
                             AND date(l.listen_date) >= date(?) AND date(l.listen_date) < date(?)), 0) AS listens,
                   COALESCE((SELECT SUM(bt.amount) FROM balance_transactions bt WHERE bt.user_id = u.id
                             AND bt.status = 'completed' AND datetime(bt.created_at) >= datetime(?)
                             AND datetime(bt.created_at) < datetime(?)), 0) AS quarter_amount
            FROM users u
            LEFT JOIN releases r ON r.user_id = u.id
            GROUP BY u.id
            ORDER BY u.id
            """,
            (start.isoformat(), end.isoformat(), start.isoformat(), end.isoformat()),
        ).fetchall()
        releases = connection.execute(
            """
            SELECT r.id, r.user_id, u.email, r.title, r.artist_name, r.release_type,
                   r.status, r.upc, r.release_date,
                   COUNT(DISTINCT rt.id) AS tracks,
                   COALESCE((SELECT SUM(l.count) FROM listens l WHERE l.release_id = r.id
                             AND date(l.listen_date) >= date(?) AND date(l.listen_date) < date(?)), 0) AS listens
            FROM releases r
            JOIN users u ON u.id = r.user_id
            LEFT JOIN release_tracks rt ON rt.release_id = r.id
            GROUP BY r.id
            ORDER BY r.id
            """,
            (start.isoformat(), end.isoformat()),
        ).fetchall()
        tracks = connection.execute(
            """
            SELECT rt.id, rt.release_id, rt.user_id, u.email, r.title AS release_title,
                   rt.title, rt.artists, rt.isrc,
                   COALESCE((SELECT SUM(tl.count) FROM track_listens tl WHERE tl.track_id = rt.id
                             AND date(tl.listen_date) >= date(?) AND date(tl.listen_date) < date(?)), 0) AS listens
            FROM release_tracks rt
            JOIN releases r ON r.id = rt.release_id
            JOIN users u ON u.id = rt.user_id
            ORDER BY rt.id
            """,
            (start.isoformat(), end.isoformat()),
        ).fetchall()

    total_listens = sum(int(row["listens"] or 0) for row in users)
    total_amount = sum(int(row["quarter_amount"] or 0) for row in users)
    summary.append(["Показатель", "Значение"])
    summary.append(["Период", f"Q{quarter} {today.year}"])
    summary.append(["Пользователей", len(users)])
    summary.append(["Релизов", len(releases)])
    summary.append(["Треков", len(tracks)])
    summary.append(["Прослушиваний", total_listens])
    summary.append(["Начислено, ₽", total_amount])

    users_sheet.append(["ID", "Имя", "Email", "Username", "Баланс", "Ожидает выплаты", "Релизов", "Прослушиваний за квартал", "Начислено за квартал"])
    for row in users:
        users_sheet.append(list(row))

    releases_sheet.append(["ID", "User ID", "Email", "Название", "Артист", "Тип", "Статус", "UPC", "Дата выхода", "Треков", "Прослушиваний за квартал"])
    for row in releases:
        releases_sheet.append(list(row))

    tracks_sheet.append(["ID", "Release ID", "User ID", "Email", "Релиз", "Трек", "Артисты", "ISRC", "Прослушиваний за квартал"])
    for row in tracks:
        tracks_sheet.append(list(row))

    _style_workbook(workbook)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _normalize_header(value: Any) -> str:
    clean = re.sub(r"\s+", " ", str(value or "").strip().lower())
    return HEADER_ALIASES.get(clean, clean.replace(" ", "_"))


def _parse_date(value: Any) -> str:
    if value in (None, ""):
        return date.today().isoformat()
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    clean = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(clean, fmt).date().isoformat()
        except ValueError:
            continue
    raise XLSXImportError("Некорректная дата. Используйте ГГГГ-ММ-ДД.")


def _parse_int(value: Any, label: str, *, allow_negative: bool = False) -> int | None:
    if value in (None, ""):
        return None
    try:
        number = int(float(str(value).replace(" ", "").replace(",", ".")))
    except (TypeError, ValueError):
        raise XLSXImportError(f"Поле «{label}» должно быть целым числом.")
    if not allow_negative and number < 0:
        raise XLSXImportError(f"Поле «{label}» не может быть отрицательным.")
    return number


async def process_import_file(
    admin_user_id: int,
    upload: UploadFile | None,
    *,
    period_label: str,
) -> tuple[bool, str, int | None]:
    if not upload or not upload.filename:
        return False, "Выберите XLSX-файл.", None
    if Path(upload.filename).suffix.lower() != ".xlsx":
        await upload.close()
        return False, "Поддерживается только формат XLSX.", None
    ensure_import_directory()
    data = await upload.read(MAX_IMPORT_BYTES + 1)
    await upload.close()
    if len(data) > MAX_IMPORT_BYTES:
        return False, "Файл слишком большой. Максимальный размер — 15 МБ.", None
    stored_name = f"{uuid.uuid4().hex}.xlsx"
    stored_path = IMPORT_DIR / stored_name
    stored_path.write_bytes(data)
    now = datetime.now(timezone.utc).isoformat()

    with _connect() as connection:
        cursor = connection.execute(
            """
            INSERT INTO analytics_imports (
                uploaded_by, original_filename, stored_filename, period_label,
                status, created_at
            ) VALUES (?, ?, ?, ?, 'processing', ?)
            """,
            (admin_user_id, Path(upload.filename).name[:180], stored_name, period_label[:80] or None, now),
        )
        import_id = int(cursor.lastrowid)
        connection.commit()

    success_count = 0
    error_count = 0
    row_count = 0
    errors: list[str] = []
    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.active
        raw_headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        headers = [_normalize_header(value) for value in raw_headers]
        if "user_email" not in headers:
            raise XLSXImportError("В первой строке должен быть столбец user_email.")
        if not any(header in headers for header in ("listens", "amount_rub")):
            raise XLSXImportError("Добавьте столбец listens или amount_rub.")

        with _connect() as connection:
            for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(value not in (None, "") for value in values):
                    continue
                row_count += 1
                data_row = {headers[index]: values[index] if index < len(values) else None for index in range(len(headers))}
                message = ""
                user_id = release_id = track_id = listens = amount = None
                status_value = "success"
                connection.execute("SAVEPOINT analytics_row")
                try:
                    email = str(data_row.get("user_email") or "").strip().lower()
                    if not email:
                        raise XLSXImportError("Не указан user_email.")
                    user = connection.execute(
                        "SELECT id, balance FROM users WHERE email = ? COLLATE NOCASE",
                        (email,),
                    ).fetchone()
                    if not user:
                        raise XLSXImportError(f"Пользователь {email} не найден.")
                    user_id = int(user["id"])
                    release_id = _parse_int(data_row.get("release_id"), "release_id")
                    track_id = _parse_int(data_row.get("track_id"), "track_id")
                    listens = _parse_int(data_row.get("listens"), "listens")
                    amount = _parse_int(data_row.get("amount_rub"), "amount_rub", allow_negative=True)
                    listen_date = _parse_date(data_row.get("date"))
                    title = " ".join(str(data_row.get("operation_title") or "Начисление по XLSX").strip().split())[:180]

                    if listens is None and amount is None:
                        raise XLSXImportError("В строке нет прослушиваний или суммы.")
                    if listens is not None:
                        if not release_id:
                            raise XLSXImportError("Для прослушиваний укажите release_id.")
                        release = connection.execute(
                            "SELECT id FROM releases WHERE id = ? AND user_id = ?",
                            (release_id, user_id),
                        ).fetchone()
                        if not release:
                            raise XLSXImportError("Релиз не найден у указанного пользователя.")
                        if track_id:
                            track = connection.execute(
                                "SELECT id FROM release_tracks WHERE id = ? AND release_id = ? AND user_id = ?",
                                (track_id, release_id, user_id),
                            ).fetchone()
                            if not track:
                                raise XLSXImportError("Трек не найден в указанном релизе.")
                        connection.execute(
                            """
                            INSERT INTO listens (user_id, release_id, listen_date, count, created_at)
                            VALUES (?, ?, ?, ?, ?)
                            """,
                            (user_id, release_id, listen_date, listens, now),
                        )
                        if track_id:
                            connection.execute(
                                """
                                INSERT INTO track_listens (
                                    user_id, release_id, track_id, listen_date, count, created_at, updated_at
                                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                                """,
                                (user_id, release_id, track_id, listen_date, listens, now, now),
                            )
                    if amount is not None and amount != 0:
                        current_balance = int(user["balance"] or 0)
                        if current_balance + amount < 0:
                            raise XLSXImportError("Списание превышает текущий баланс пользователя.")
                        connection.execute(
                            "UPDATE users SET balance = balance + ?, updated_at = ? WHERE id = ?",
                            (amount, now, user_id),
                        )
                        connection.execute(
                            """
                            INSERT INTO balance_transactions (
                                user_id, release_id, transaction_type, status, amount, title, created_at
                            ) VALUES (?, ?, 'xlsx_import', 'completed', ?, ?, ?)
                            """,
                            (user_id, release_id, amount, title or "Начисление по XLSX", now),
                        )
                    message = "Строка импортирована."
                    connection.execute("RELEASE SAVEPOINT analytics_row")
                    success_count += 1
                except XLSXImportError as exc:
                    connection.execute("ROLLBACK TO SAVEPOINT analytics_row")
                    connection.execute("RELEASE SAVEPOINT analytics_row")
                    status_value = "error"
                    message = str(exc)
                    error_count += 1
                    errors.append(f"Строка {row_number}: {message}")
                except sqlite3.Error as exc:
                    try:
                        connection.execute("ROLLBACK TO SAVEPOINT analytics_row")
                        connection.execute("RELEASE SAVEPOINT analytics_row")
                    except sqlite3.Error:
                        pass
                    status_value = "error"
                    message = "Ошибка базы данных при обработке строки."
                    error_count += 1
                    errors.append(f"Строка {row_number}: {exc}")
                connection.execute(
                    """
                    INSERT INTO analytics_import_rows (
                        import_id, row_number, status, message, user_id, release_id,
                        track_id, listens, amount, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (import_id, row_number, status_value, message, user_id, release_id, track_id, listens, amount, now),
                )
            status_final = "completed" if error_count == 0 else "partial" if success_count else "failed"
            connection.execute(
                """
                UPDATE analytics_imports
                SET status = ?, row_count = ?, success_count = ?, error_count = ?,
                    error_summary = ?, completed_at = ?
                WHERE id = ?
                """,
                (status_final, row_count, success_count, error_count, "\n".join(errors[:30]) or None, datetime.now(timezone.utc).isoformat(), import_id),
            )
            _audit(
                connection,
                admin_user_id,
                "xlsx_import",
                "analytics_import",
                import_id,
                {"rows": row_count, "success": success_count, "errors": error_count},
            )
            connection.commit()
    except (XLSXImportError, OSError, ValueError) as exc:
        with _connect() as connection:
            connection.execute(
                """
                UPDATE analytics_imports
                SET status = 'failed', error_summary = ?, completed_at = ?
                WHERE id = ?
                """,
                (str(exc), datetime.now(timezone.utc).isoformat(), import_id),
            )
            connection.commit()
        return False, str(exc), import_id

    if success_count and error_count:
        return True, f"Импортировано строк: {success_count}. Пропущено с ошибками: {error_count}.", import_id
    if success_count:
        return True, f"Успешно импортировано строк: {success_count}.", import_id
    return False, "Не удалось импортировать ни одной строки. Проверьте файл.", import_id


def resolve_import_file(stored_filename: str | None) -> Path | None:
    if not stored_filename:
        return None
    ensure_import_directory()
    candidate = (IMPORT_DIR / Path(stored_filename).name).resolve()
    try:
        candidate.relative_to(IMPORT_DIR.resolve())
    except ValueError:
        return None
    return candidate if candidate.is_file() else None
